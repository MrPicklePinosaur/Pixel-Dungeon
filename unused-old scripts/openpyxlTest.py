import openpyxl as o

#load workbook and spreadsheet
wb = o.load_workbook(filename = 'testSheet.xlsx')
sheet1 = wb['Sheet1']


#Get the number of rows and columns
rowCount = sheet1.max_row
columnCount = sheet1.max_column

statDict = {} #Dict to organize spreadsheet values

#Iterate through spreadsheet
for i in range(2,rowCount+1): #For each row (besides the key type row [1])
    statDict[sheet1.cell(row=i, column=1).value] = {} #Create a new dict for th stats for each item
    for n in range(2,columnCount+1): #For each column (besides the key name column [A])
        cellValue = sheet1.cell(row=i, column=n).value #Get value at specific cell
        #add data to dict
        statDict[sheet1.cell(row=i, column=1).value][sheet1.cell(row=1, column=n).value] = cellValue

print(statDict)



#print(sheet1['A3'].value)
