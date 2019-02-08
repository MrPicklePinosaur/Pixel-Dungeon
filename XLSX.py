import openpyxl as o

class XLSX:

    @staticmethod
    def load(workbook,page): #Load data from spreadsheet and put it into a dict
        #load workbook and spreadsheet
        wb = o.load_workbook(filename = workbook)
        sheet = wb[page]

        #Get the number of rows and columns of sheet
        rowCount = sheet.max_row
        columnCount = sheet.max_column

        dataDict = {} #Dict to organize spreadsheet values

        #Iterate through spreadsheet
        for i in range(2,rowCount+1): #For each row (besides the key type row [1])
            dataDict[sheet.cell(row=i, column=1).value] = {} #Create a new dict for th stats for each item
            for n in range(1,columnCount+1): #For each column 
                cellValue = sheet.cell(row=i, column=n).value #Get value at specific cell
                '''
                #If the value is a number convert it into an int object
                if cellValue.isdigit() == True:
                    cellValue = int(cellValue)
                '''
                #Add data to dict
                dataDict[sheet.cell(row=i, column=1).value][sheet.cell(row=1, column=n).value] = cellValue

        return dataDict

    @staticmethod
    #Takes in a dict of items and stats, and creates those and stores them in a list
    def createObjects(itemType,staticDict):
        #For each object in dict, create an actual instance with the stats of the dict object
        #Note that the name of the key is the first parameter
        objectDict = {staticDict[n]['name']:itemType(*(staticDict[n][i] for i in staticDict[n])) for n in staticDict}
        return objectDict
    
#print(XLSX.load('assets/data/weaponStats.xlsx','melee'))
